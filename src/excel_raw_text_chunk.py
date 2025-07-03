from openpyxl import load_workbook
import json

def extract_employee_chunks(file_path: str, output_path: str = "employee_chunks_raw.json"):
    # === Load the workbook ===
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active

    # === Read all rows as raw strings ===
    rows = []
    for row in sheet.iter_rows(values_only=True):
        raw_line = "\t".join(str(cell).strip() if cell else "" for cell in row).strip()
        rows.append(raw_line)

    # === Group rows into chunks per employee (starts with 'Emp#') ===
    employee_chunks = []
    current_chunk = []

    for line in rows:
        if "Emp#" in line:
            if current_chunk:
                employee_chunks.append("\n".join(current_chunk))
                current_chunk = []
        current_chunk.append(line)

    # Add the final chunk
    if current_chunk:
        employee_chunks.append("\n".join(current_chunk))

    # === Save to JSON ===
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(employee_chunks, f, indent=2)

    print(f"✅ Done. Extracted {len(employee_chunks)} employee chunks.")
